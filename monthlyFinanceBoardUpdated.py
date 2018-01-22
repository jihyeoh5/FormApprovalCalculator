import os
import openpyxl
import sys
import re
from openpyxl.utils import column_index_from_string

def changeDate(sheet,dateObject): #Change the date formatting: Iterates through the row/columns and change the value of the cell to first 10 characters if it matches the pattern dateObject
        for rowNum in range(1,sheet.max_row+1):
                for colNum in range(1,sheet.max_column):
                        unchangedDate = str(sheet.cell(row=rowNum,column=colNum).value)
                        if dateObject.search(unchangedDate): sheet.cell(row=rowNum,column=colNum).value = unchangedDate[:10]
def calculateManagerDifference(sheet,wb,emptyRows,listSheets,dictOutput,form): #Calculates difference for Manager
        type = 'Manager'
        managerAverage = 0
        managerSheet= wb.create_sheet('Manager')
        counter = 0 #Counts how many rows to skip
        #Making Manager Sheet by copying data. Skips rows that has empty value for 'Approved'
        for rowNum in range(1,sheet.max_row+1):
                if rowNum not in emptyRows:
                        for colNum in range(1, sheet.max_column): managerSheet.cell(row=rowNum-counter,column=colNum).value = sheet.cell(row=rowNum,column=colNum).value
                else:
                        for colNum in range(1, sheet.max_column): managerSheet.cell(row=rowNum-counter,column=colNum).value = sheet.cell(row=rowNum+1,column=colNum).value
                        counter = counter+1
        #Adding Calculated Manager Column
        calculatedManager=sheet.max_column+1
        managerSheet.cell(row=1,column=calculatedManager).value= "Calculated"
        #Find column number of "Modified" and "Approved"
        for colNum in range(1,5):
                if wb[listSheets[0]].cell(row=1,column=colNum).value == "Created": createdColNum = colNum
                elif wb[listSheets[0]].cell(row=1,column=colNum).value == "Approval Date" or wb[listSheets[0]].cell(row=1,column=colNum).value == "Approved Date" or wb[listSheets[0]].cell(row=1,column=colNum).value=="Approved By Date":
                        approvedColNum = colNum
        #Writing in the formula for manager calculation
        for rowNum in range(2,sheet.max_row - counter):
                calculated = int(managerSheet.cell(row=rowNum,column=int(approvedColNum)).value[8:10])-int(managerSheet.cell(row=rowNum,column=int(createdColNum)).value[8:10])
                numDays = [31,30,31,30,31,30,31,31,30,31,30,31]
                month=int(managerSheet.cell(row=rowNum, column=int(approvedColNum)).value[5:7])
                if calculated < 0:
                        calculated = int(managerSheet.cell(row=rowNum, column=int(approvedColNum)).value[8:10])+(numDays[month-1]-int(managerSheet.cell(row=rowNum, column=int(createdColNum)).value[8:10]))
                managerSheet.cell(row=rowNum,column=calculatedManager).value= calculated
                managerAverage = managerAverage + calculated
        dictOutput[form]["manager"]=findAverage(managerSheet,type,managerAverage) #Update dictionary with managerAverage
def calculateFinanceDifference(sheet,inNumbers,wb,emptyRows,listSheets,dictOutput,form): #Calculates difference for Finance
        financeAverage = 0
        type = 'Finance'
        financeSheet= wb.create_sheet('Finance')
        #Making Finance Sheet. Copies over data. If Approved cell is empty, then replace it with value from 'Created' column
        for rowNum in range(1,sheet.max_row+1):
                if rowNum in emptyRows:
                        for colNum in range(1, sheet.max_column):
                                financeSheet.cell(row=rowNum,column=colNum).value = sheet.cell(row=rowNum,column=colNum).value
                                financeSheet.cell(row=rowNum,column=inNumbers[2]).value = sheet.cell(row=rowNum,column=inNumbers[0]).value
                else:
                        for colNum in range(1, sheet.max_column): financeSheet.cell(row=rowNum,column=colNum).value = sheet.cell(row=rowNum,column=colNum).value
        #Adding Calculated Manager Column
        calculatedFinance=sheet.max_column+1
        financeSheet.cell(row=1,column=calculatedFinance).value= "Calculated"
        #Find column number of "Modified" and "Approved"
        for colNum in range(1,5):
                if wb[listSheets[0]].cell(row=1,column=colNum).value == "Modified": modifiedColNum = colNum
                elif wb[listSheets[0]].cell(row=1,column=colNum).value == "Approval Date" or wb[listSheets[0]].cell(row=1,column=colNum).value == "Approved Date" or wb[listSheets[0]].cell(row=1,column=colNum).value == "Approved By Date":
                        approvedColNum = colNum
        #Doing the actual calculation
        for rowNum in range(2,sheet.max_row):
                calculated = int(financeSheet.cell(row=rowNum,column=int(modifiedColNum)).value[8:10])-int(financeSheet.cell(row=rowNum,column=int(approvedColNum)).value[8:10])
                numDays = [31,30,31,30,31,30,31,31,30,31,30,31]
                month=int(financeSheet.cell(row=rowNum, column=int(approvedColNum)).value[5:7])
                if calculated < 0:
                        calculated = int(financeSheet.cell(row=rowNum, column=int(modifiedColNum)).value[8:10])+(numDays[month-1]-int(financeSheet.cell(row=rowNum, column=int(approvedColNum)).value[8:10]))
                financeSheet.cell(row=rowNum,column=calculatedFinance).value=calculated
                financeAverage = financeAverage + calculated
        dictOutput[form]["finance"]=findAverage(financeSheet,type,financeAverage)
def findColumnsAndEmptyRows(sheet,emptyRows,inNumbers):
#Find the last column & also find the rows that have empty values - for different calculations
        #Iterate through columns and find matching values.
        for colNum in range(1,sheet.max_column+1):
                search = str(sheet.cell(row=1,column=colNum).value)
                if search == "Created":
                        emptyRows[0] = str(sheet.cell(row=1,column=colNum).column)
                        inNumbers[0] = column_index_from_string(emptyRows[0])
                elif search == "Modified":
                        emptyRows[1] = str(sheet.cell(row=1,column=colNum).column)
                        inNumbers[1] = column_index_from_string(emptyRows[1])
                elif search == "Approved By Date":
                        emptyRows[2] = str(sheet.cell(row=1,column=colNum).column)
                        inNumbers[2] = column_index_from_string(emptyRows[2])
                elif search == "Approval Date" or search == "Approved Date":
                        emptyRows[2] = str(sheet.cell(row=1,column=colNum).column)
                        inNumbers[2] = column_index_from_string(emptyRows[2])
        #Find rows where Approved Date is blank
        for rowNum in range(1, sheet.max_row+1):
                if sheet.cell(row=rowNum,column=inNumbers[2]).value == None: emptyRows.append(sheet.cell(row=rowNum,column=inNumbers[2]).row)
def findAverage(sheetname,type,average):
#Find the average and output it as a cell
        actualAverage = average/(int(sheetname.max_row)-2)
        sheetname.cell(row=2,column=sheetname.max_column+1).value= actualAverage
        sheetname.cell(row=1,column=sheetname.max_column).value= 'Average'
        actualAverage = float("{0:.2f}".format(actualAverage))
        return actualAverage
def whichForm(formName):
#Determine the name of the form
        EXsubstring = "ex" in formName
        VPsubstring = "vp" in formName
        PRsubstring = "pr" in formName
        CSPsubstring = "csp" in formName
        if EXsubstring == True: form = "Expense"
        elif VPsubstring == True: form = "Vendor Payment"
        elif PRsubstring == True: form = "Payment Requisition"
        elif CSPsubstring == True: form = "CS Payment"
        return form
def MFBscript(location):
        os.chdir(location)
        listForms = os.listdir(location)
        dictOutput = {"CS Payment":{"manager":0,"finance":0,"numForms":0},
                      "Expense":{"manager":0,"finance":0,"numForms":0},
                      "Vendor Payment":{"manager":0,"finance":0,"numForms":0},
                      "Payment Requisition":{"manager":0,"finance":0,"numForms":0},}
        for i in range(0,len(listForms)):
                emptyRows=['Created Column Location','Modified Column Location','Approved Column Location']
                inNumbers=[1,2,3]
                dateObject = re.compile(r'\d\d\d\d-\d\d-\d\d')
                loadForm = location+"\\"+listForms[i]
                wb = openpyxl.load_workbook(loadForm)
                listSheets = wb.get_sheet_names()
                firstSheet = wb[listSheets[0]]
                sheet = wb.active
                form = whichForm(listForms[i])
                changeDate(sheet,dateObject)
                findColumnsAndEmptyRows(sheet,emptyRows,inNumbers)
                calculateManagerDifference(sheet,wb, emptyRows,listSheets,dictOutput,form)
                calculateFinanceDifference(sheet, inNumbers,wb,emptyRows,listSheets,dictOutput,form)
                dictOutput[form]["numForms"]=str(sheet.max_row)
                wb.save(loadForm)

        return dictOutput
